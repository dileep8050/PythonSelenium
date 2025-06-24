import os
from openpyxl import load_workbook




def get_excel_test_data(sheet_name):
    current_dir = os.path.dirname(os.path.abspath(__file__))
    path = os.path.join(current_dir, "..", "testdata", "data.xlsx")
    file_path = os.path.abspath(path)  # Normalize the path
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]

    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        data.append(row_dict)

    return data